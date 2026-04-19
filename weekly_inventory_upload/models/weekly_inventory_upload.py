import base64
import io
import re
from collections import defaultdict

from openpyxl import Workbook, load_workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError


class WeeklyInventoryUpload(models.Model):
    _name = 'weekly.inventory.upload'
    _description = 'Weekly Inventory Upload'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'id desc'
    _template_max_row = 1048576

    name = fields.Char(
        string='Reference',
        required=True,
        copy=False,
        readonly=True,
        default=lambda self: _('New'),
        tracking=True,
    )
    upload_file = fields.Binary(string='Excel File', attachment=True)
    upload_filename = fields.Char(string='File Name')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processed', 'Processed'),
        ('undone', 'Undone'),
    ], default='draft', tracking=True)
    upload_date = fields.Datetime(string='Upload Date', readonly=True)
    uploaded_by = fields.Many2one('res.users', string='Uploaded By', readonly=True)
    total_rows = fields.Integer(string='Total Rows', readonly=True)
    success_rows = fields.Integer(string='Success Rows', readonly=True)
    failed_rows = fields.Integer(string='Failed Rows', readonly=True)
    message_summary = fields.Text(string='Summary', readonly=True)
    error_file = fields.Binary(string='Error Report', attachment=True, readonly=True)
    error_filename = fields.Char(string='Error File Name', readonly=True)
    line_ids = fields.One2many('weekly.inventory.upload.line', 'upload_id', string='Upload Logs', readonly=True)
    can_undo = fields.Boolean(compute='_compute_can_undo')

    @api.depends('state')
    def _compute_can_undo(self):
        for rec in self:
            rec.can_undo = rec.state == 'processed'

    @api.model_create_multi
    def create(self, vals_list):
        seq = self.env['ir.sequence']
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = seq.next_by_code('weekly.inventory.upload') or _('New')
        return super().create(vals_list)

    def _sanitize_download_filename(self, filename, default_name):
        safe_filename = re.sub(r'[^A-Za-z0-9._-]+', '_', (filename or '').strip()).strip('._')
        return safe_filename or default_name

    def _get_binary_download_url(self, field_name, filename_field):
        self.ensure_one()
        return '/web/content?model=%s&id=%s&field=%s&filename_field=%s&download=true' % (
            self._name,
            self.id,
            field_name,
            filename_field,
        )

    def _get_product_reference_value(self, product):
        return product.default_code or 'ID:%s' % product.id

    def _get_product_display_name(self, product):
        return product.with_context(display_default_code=False).display_name

    def _normalize_text(self, value):
        return str(value).strip() if value is not None else ''

    def _find_product_by_reference(self, product_reference):
        product_reference = (product_reference or '').strip()
        if not product_reference:
            return self.env['product.product']

        if product_reference.upper().startswith('ID:'):
            product_id_text = product_reference.split(':', 1)[1].strip()
            if product_id_text.isdigit():
                product = self.env['product.product'].search([
                    ('id', '=', int(product_id_text)),
                    ('is_storable', '=', True),
                ], limit=1)
                if product:
                    return product
            raise ValidationError(_('Product not found for reference: %s') % product_reference)

        product = self.env['product.product'].search([
            ('default_code', '=', product_reference),
            ('is_storable', '=', True),
        ], limit=2)
        if len(product) > 1:
            raise ValidationError(_(
                'Multiple products found for internal reference "%s".'
            ) % product_reference)
        if product:
            return product

        raise ValidationError(_('Product not found for reference: %s') % product_reference)

    def _find_product_from_row(self, product_reference, product_name):
        if product_reference:
            return self._find_product_by_reference(product_reference)
        if product_name:
            return self._find_product(product_name)
        raise ValidationError(_('Product reference and product are missing.'))

    def _find_product(self, product_name):
        product_name = (product_name or '').strip()

        if product_name.startswith('[ID:') and '] ' in product_name:
            product_id_text = product_name[4:].split('] ', 1)[0].strip()
            if product_id_text.isdigit():
                product = self.env['product.product'].search([
                    ('id', '=', int(product_id_text)),
                    ('is_storable', '=', True),
                ], limit=1)
                if product:
                    return product

        if product_name.startswith('[') and '] ' in product_name:
            default_code = product_name[1:].split('] ', 1)[0].strip()
            if default_code:
                product = self.env['product.product'].search([
                    ('default_code', '=', default_code),
                    ('is_storable', '=', True),
                ], limit=2)
                if len(product) > 1:
                    raise ValidationError(_(
                        'Multiple products found for internal reference "%s".'
                    ) % default_code)
                if product:
                    return product

        if ' (' in product_name and product_name.endswith(')'):
            template_name = product_name.rsplit(' (', 1)[0].strip()
            candidate_products = self.env['product.product'].search([
                ('name', '=', template_name),
                ('is_storable', '=', True),
            ])
            display_products = candidate_products.with_context(display_default_code=False).filtered(
                lambda p: p.display_name == product_name
            )
            if len(display_products) > 1:
                raise ValidationError(_(
                    'Multiple variants found for "%s". Please use the dropdown value from the sample template.'
                ) % product_name)
            if display_products:
                return display_products

        product = self.env['product.product'].search([
            ('name', '=', product_name),
            ('is_storable', '=', True),
        ], limit=2)
        if len(product) > 1:
            raise ValidationError(_(
                'Multiple products found for "%s". Please use a unique product name or internal reference.'
            ) % product_name)
        if product:
            return product

        product = self.env['product.product'].search([
            ('default_code', '=', product_name),
            ('is_storable', '=', True),
        ], limit=2)
        if len(product) > 1:
            raise ValidationError(_(
                'Multiple products found for internal reference "%s".'
            ) % product_name)
        if product:
            return product

        raise ValidationError(_('Product not found: %s') % product_name)

    def _find_location(self, location_name):
        location_domain = [('usage', 'in', ('internal', 'transit'))]
        location = self.env['stock.location'].search(
            location_domain + [('complete_name', '=', location_name)],
            limit=2,
        )
        if len(location) > 1:
            raise ValidationError(_(
                'Multiple locations found for "%s". Please use a unique full location name.'
            ) % location_name)
        if location:
            return location

        location = self.env['stock.location'].search(
            location_domain + [('barcode', '=', location_name)],
            limit=2,
        )
        if len(location) > 1:
            raise ValidationError(_(
                'Multiple locations found for barcode "%s".'
            ) % location_name)
        if location:
            return location

        location = self.env['stock.location'].search(
            location_domain + [('name', '=', location_name)],
            limit=2,
        )
        if len(location) > 1:
            raise ValidationError(_(
                'Multiple locations found for "%s". Please use the full location name.'
            ) % location_name)
        if location:
            return location

        raise ValidationError(_('Location not found: %s') % location_name)

    def _get_template_products(self):
        return self.env['product.product'].search(
            [('is_storable', '=', True)],
            order='default_code, name, id',
        )

    def _get_template_locations(self):
        return self.env['stock.location'].search(
            [('usage', 'in', ('internal', 'transit'))],
            order='complete_name, id',
        )

    def _resolve_location_headers(self, headers):
        location_by_header = {}
        for header in headers[2:]:
            location_name = self._normalize_text(header)
            if not location_name or location_name in location_by_header:
                continue
            location_by_header[location_name] = self._find_location(location_name)
        return location_by_header

    def _prepare_wide_format_entries(self, rows, headers):
        location_by_header = self._resolve_location_headers(headers)
        if not location_by_header:
            raise ValidationError(_('The template must include at least one location column after Product Reference and Product.'))

        total_rows = 0
        error_rows = []
        grouped_rows = defaultdict(list)

        for row_index, row in enumerate(rows[1:], start=2):
            product_reference = self._normalize_text(row[0] if len(row) > 0 else '')
            product_name = self._normalize_text(row[1] if len(row) > 1 else '')
            location_cells = row[2:] if len(row) > 2 else []

            if not product_reference and not product_name and all(cell in (None, '') for cell in location_cells):
                continue

            for column_index, header in enumerate(headers[2:], start=2):
                location_name = self._normalize_text(header)
                if not location_name:
                    continue
                cell_value = row[column_index] if column_index < len(row) else None
                if cell_value in (None, ''):
                    continue
                if isinstance(cell_value, str) and not cell_value.strip():
                    continue

                try:
                    quantity_value = float(cell_value)
                except Exception:
                    total_rows += 1
                    error_rows.append([
                        row_index,
                        product_reference,
                        product_name,
                        location_name,
                        cell_value,
                        _('Quantity must be numeric.'),
                    ])
                    continue

                if quantity_value < 0:
                    total_rows += 1
                    error_rows.append([
                        row_index,
                        product_reference,
                        product_name,
                        location_name,
                        quantity_value,
                        _('Quantity cannot be negative.'),
                    ])
                    continue

                if quantity_value == 0:
                    continue

                total_rows += 1
                grouped_rows[(product_reference, product_name, location_name)].append({
                    'row_number': row_index,
                    'product_reference': product_reference,
                    'product_name': product_name,
                    'location_name': location_name,
                    'quantity': quantity_value,
                    'location': location_by_header[location_name],
                })

        return total_rows, grouped_rows, error_rows

    def _prepare_long_format_entries(self, rows, headers):
        valid_headers = [
            ['Product', 'Location', 'Quantity'],
            ['Product', 'Warehouse Location', 'Quantity'],
            ['Product', 'Warehouse', 'Quantity'],
        ]
        if headers[:3] not in valid_headers:
            raise ValidationError(_(
                'Invalid template. Use either Product Reference, Product, then location columns; '
                'or Product, Location, Quantity.'
            ))

        total_rows = 0
        error_rows = []
        grouped_rows = defaultdict(list)

        for row_index, row in enumerate(rows[1:], start=2):
            if not row or all(cell in (None, '') for cell in row[:3]):
                continue

            product_name = self._normalize_text(row[0] if len(row) > 0 else '')
            location_name = self._normalize_text(row[1] if len(row) > 1 else '')
            quantity_value = row[2] if len(row) > 2 else None

            if quantity_value in (None, '') or (isinstance(quantity_value, str) and not quantity_value.strip()):
                total_rows += 1
                error_rows.append([
                    row_index,
                    '',
                    product_name,
                    location_name,
                    '',
                    _('Quantity is missing.'),
                ])
                continue

            try:
                quantity_value = float(quantity_value)
            except Exception:
                total_rows += 1
                error_rows.append([
                    row_index,
                    '',
                    product_name,
                    location_name,
                    quantity_value,
                    _('Quantity must be numeric.'),
                ])
                continue

            if quantity_value < 0:
                total_rows += 1
                error_rows.append([
                    row_index,
                    '',
                    product_name,
                    location_name,
                    quantity_value,
                    _('Quantity cannot be negative.'),
                ])
                continue

            if quantity_value == 0:
                continue

            total_rows += 1
            grouped_rows[('', product_name, location_name)].append({
                'row_number': row_index,
                'product_reference': '',
                'product_name': product_name,
                'location_name': location_name,
                'quantity': quantity_value,
                'location': False,
            })

        return total_rows, grouped_rows, error_rows

    def action_download_sample_template(self):
        self.ensure_one()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Inventory Upload'
        locations = self._get_template_locations()
        products = self._get_template_products()

        headers = ['Product Reference', 'Product'] + locations.mapped('complete_name')
        sheet.append(headers)

        for product in products:
            sheet.append([
                self._get_product_reference_value(product),
                self._get_product_display_name(product),
            ] + [''] * len(locations))

        sheet.freeze_panes = 'C2'
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40
        for column_index in range(3, len(headers) + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = 18

        output = io.BytesIO()
        workbook.save(output)
        file_content = base64.b64encode(output.getvalue())
        self.write({
            'error_file': file_content,
            'error_filename': self._sanitize_download_filename(
                'sample_weekly_inventory_template.xlsx',
                'sample_weekly_inventory_template.xlsx',
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': self._get_binary_download_url('error_file', 'error_filename'),
            'target': 'self',
        }

    def action_process_upload(self):
        self.ensure_one()
        if self.state != 'draft':
            raise UserError(_('Only draft uploads can be processed.'))
        if not self.upload_file:
            raise UserError(_('Please upload an Excel file first.'))

        self.line_ids.unlink()
        self.write({
            'error_file': False,
            'error_filename': False,
            'message_summary': False,
            'total_rows': 0,
            'success_rows': 0,
            'failed_rows': 0,
        })

        file_data = base64.b64decode(self.upload_file)
        workbook = load_workbook(filename=io.BytesIO(file_data), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('The uploaded file is empty.'))

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        is_wide_format = len(headers) >= 3 and headers[0] == 'Product Reference' and headers[1] == 'Product'
        if is_wide_format:
            total_rows, grouped_rows, error_rows = self._prepare_wide_format_entries(rows, headers)
        else:
            total_rows, grouped_rows, error_rows = self._prepare_long_format_entries(rows, headers)

        success_rows = 0
        failed_rows = len(error_rows)

        for grouped in grouped_rows.values():
            last_entry = grouped[-1]
            product_reference = last_entry['product_reference']
            product_name = last_entry['product_name']
            location_name = last_entry['location_name']
            final_qty = last_entry['quantity']
            last_row_no = last_entry['row_number']
            duplicate_rows = [entry['row_number'] for entry in grouped]
            duplicate_note = ''
            if len(duplicate_rows) > 1:
                duplicate_note = _(' Duplicate rows detected at Excel rows: %s. Last row quantity used.') % ', '.join(map(str, duplicate_rows))

            try:
                if not product_reference and not product_name:
                    raise ValidationError(_('Product reference and product are missing.'))
                if not location_name:
                    raise ValidationError(_('Location is missing.'))

                product = self._find_product_from_row(product_reference, product_name)
                location = last_entry['location'] or self._find_location(location_name)
                quant = self.env['stock.quant'].search([
                    ('product_id', '=', product.id),
                    ('location_id', '=', location.id),
                ], limit=1)
                if not quant:
                    raise ValidationError(_(
                        'Product "%(product)s" is not available at location "%(location)s". Quantity was not updated.',
                        product=product.display_name,
                        location=location.complete_name,
                    ))

                old_qty = quant.quantity

                quant.inventory_quantity = final_qty
                quant._apply_inventory()

                self.env['weekly.inventory.upload.line'].create({
                    'upload_id': self.id,
                    'row_number': last_row_no,
                    'product_reference': product_reference or self._get_product_reference_value(product),
                    'product_name': self._get_product_display_name(product),
                    'warehouse_name': location_name,
                    'product_id': product.id,
                    'warehouse_id': location.warehouse_id.id,
                    'location_id': location.id,
                    'old_qty': old_qty,
                    'new_qty': final_qty,
                    'status': 'success',
                    'message': _('Inventory updated successfully.') + duplicate_note,
                })
                success_rows += 1
            except Exception as exc:
                failed_rows += 1
                message = str(exc)
                error_rows.append([
                    last_row_no,
                    product_reference,
                    product_name,
                    location_name,
                    final_qty if final_qty not in (None, '') else '',
                    message,
                ])
                self.env['weekly.inventory.upload.line'].create({
                    'upload_id': self.id,
                    'row_number': last_row_no,
                    'product_reference': product_reference,
                    'product_name': product_name,
                    'warehouse_name': location_name,
                    'new_qty': final_qty if isinstance(final_qty, (int, float)) else 0.0,
                    'status': 'failed',
                    'message': message,
                })

        summary = _('Inventory updated successfully for %s entries') % success_rows
        if failed_rows:
            summary += _('. Failed rows: %s') % failed_rows
            self._generate_error_report(error_rows)

        self.write({
            'state': 'processed',
            'upload_date': fields.Datetime.now(),
            'uploaded_by': self.env.user.id,
            'total_rows': total_rows,
            'success_rows': success_rows,
            'failed_rows': failed_rows,
            'message_summary': summary,
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Upload Complete'),
                'message': summary,
                'sticky': False,
                'type': 'success' if not failed_rows else 'warning',
                'next': {
                    'type': 'ir.actions.client',
                    'tag': 'reload',
                }
            }
        }

    def _generate_error_report(self, error_rows):
        self.ensure_one()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Error Report'
        sheet.append(['Excel Row', 'Product Reference', 'Product', 'Location', 'Quantity', 'Error Message'])
        for row in error_rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        self.write({
            'error_file': base64.b64encode(output.getvalue()),
            'error_filename': self._sanitize_download_filename(
                '%s_error_report.xlsx' % self.name,
                'error_report.xlsx',
            ),
        })

    def action_download_error_report(self):
        self.ensure_one()
        if not self.error_file:
            raise UserError(_('No error report found for this upload.'))
        safe_filename = self._sanitize_download_filename(self.error_filename, 'error_report.xlsx')
        if self.error_filename != safe_filename:
            self.error_filename = safe_filename
        return {
            'type': 'ir.actions.act_url',
            'url': self._get_binary_download_url('error_file', 'error_filename'),
            'target': 'self',
        }

    def action_undo_upload(self):
        self.ensure_one()
        if self.state != 'processed':
            raise UserError(_('Only processed uploads can be undone.'))

        success_lines = self.line_ids.filtered(lambda l: l.status == 'success')
        if not success_lines:
            raise UserError(_('No successful lines found to undo.'))

        for line in success_lines:
            quant = self.env['stock.quant'].search([
                ('product_id', '=', line.product_id.id),
                ('location_id', '=', line.location_id.id),
            ], limit=1)
            if quant:
                quant.inventory_quantity = line.old_qty
                quant._apply_inventory()
            else:
                new_quant = self.env['stock.quant'].with_context(inventory_mode=True).create({
                    'product_id': line.product_id.id,
                    'location_id': line.location_id.id,
                    'inventory_quantity': line.old_qty,
                })
                new_quant._apply_inventory()
            line.undo_done = True

        self.write({
            'state': 'undone',
            'message_summary': _('Inventory restored successfully.'),
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Undo Complete'),
                'message': _('Inventory restored successfully.'),
                'sticky': False,
                'type': 'success',
                 'next': {
                    'type': 'ir.actions.client',
                    'tag': 'reload',
                }
            }
        }


class WeeklyInventoryUploadLine(models.Model):
    _name = 'weekly.inventory.upload.line'
    _description = 'Weekly Inventory Upload Line'
    _order = 'id asc'

    upload_id = fields.Many2one('weekly.inventory.upload', string='Upload', required=True, ondelete='cascade')
    row_number = fields.Integer(string='Excel Row')
    product_reference = fields.Char(string='Product Reference')
    product_name = fields.Char(string='Product')
    warehouse_name = fields.Char(string='Location')
    product_id = fields.Many2one('product.product', string='Matched Product')
    warehouse_id = fields.Many2one('stock.warehouse', string='Related Warehouse')
    location_id = fields.Many2one('stock.location', string='Matched Location')
    old_qty = fields.Float(string='Old Quantity', digits='Product Unit of Measure')
    new_qty = fields.Float(string='New Quantity', digits='Product Unit of Measure')
    status = fields.Selection([
        ('success', 'Success'),
        ('failed', 'Failed'),
    ], string='Status', default='success')
    message = fields.Text(string='Message')
    undo_done = fields.Boolean(string='Undo Done', default=False)
