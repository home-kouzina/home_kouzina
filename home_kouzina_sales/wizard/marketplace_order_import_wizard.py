import base64
import io
from datetime import datetime
from odoo import models, fields, api
from odoo.exceptions import UserError
import xlsxwriter
from openpyxl import load_workbook


class MarketplaceImportLog(models.TransientModel):
    _name = "marketplace.import.log"
    _description = "Marketplace Import Log"
    _rec_name = 'wizard_id'

    wizard_id = fields.Many2one('marketplace.order.import.wizard', string='Wizard')
    row_index = fields.Integer(string='Row #')
    status = fields.Selection([('success', 'Success'), ('failed', 'Failed'), ('warning', 'Warning')], string='Status')
    message = fields.Text(string='Message')


class MarketplaceOrderImportWizard(models.TransientModel):
    _name = 'marketplace.order.import.wizard'
    _description = 'Import Marketplace Orders from XLSX'
    _rec_name = 'marketplace_id'

    # Optional wizard field; allows multiple mixed marketplaces in a single sheet
    marketplace_id = fields.Many2one('marketplace.master', string='Marketplace', required=False)
    xlsx_file = fields.Binary(string='Upload XLSX File', required=True)
    xlsx_filename = fields.Char(string='Filename')
    confirm_orders = fields.Boolean(string='Confirm Created Orders', default=False)
    create_customer_if_missing = fields.Boolean(string='Create Customers if Missing', default=True)
    log_ids = fields.One2many('marketplace.import.log', 'wizard_id', string='Import Log')
    report_file = fields.Binary(string='Failed Orders Report')
    report_filename = fields.Char(string='Report Filename')

    DEFAULT_EXPECTED_HEADERS = [
        'marketplace_order_id', 'marketplace_type_excel', 'customer_email', 'customer_name',
        'customer_street', 'customer_city', 'customer_zip',
        'order_date', 'product_sku', 'quantity', 'unit_price',
        'order_line_note', 'marketplace_invoice_number',
        'marketplace_invoice_type', 'marketplace_sale_state', 'marketplace_hsn_code'
    ]

    HEADER_MAP = {
        'Marketplace Type': 'marketplace_type_excel',
        'Marketplace Order ID': 'marketplace_order_id',
        'Supply City': 'supply_city',
        'Supply State': 'supply_state',
        'Supply State GST': 'supply_state_gst',
        'GSTIN': 'vat',
        'GST Number': 'vat',
        'Customer GSTIN': 'vat',
        'HSN Code': 'marketplace_hsn_code',
        'Marketplace HSN Code': 'marketplace_hsn_code',
        'Item ID': 'marketplace_item_id',
        'Customer Email': 'customer_email',
        'Customer Name': 'customer_name',
        'Customer Street': 'customer_street',
        'Customer City': 'customer_city',
        'Customer State': 'customer_state',
        'Customer Zip': 'customer_zip',
        'Customer Country': 'customer_country',
        'Billing Street': 'customer_street',
        'Billing City': 'customer_city',
        'Billing State': 'customer_state',
        'Billing Zip': 'customer_zip',
        'Billing Country': 'customer_country',
        'Order Date': 'order_date',
        'Product SKU': 'product_sku',
        'Product Name': 'product_name',
        'Quantity': 'quantity',
        'Unit Price': 'unit_price',
        'Order Line Note': 'order_line_note',
        'Tax Percent': 'tax_percent',
        'Shipping Amount': 'shipping_amount',
        'Currency': 'currency',
        'Order Tag': 'order_tag',
        'Order ID': 'marketplace_order_ref',
        'Marketplace Order Date': 'marketplace_order_date',
        'Payment Status': 'marketplace_payment_status',
        'Order Status': 'marketplace_order_status',
        'Delivery Slot': 'marketplace_delivery_slot',
        'Total Amount': 'marketplace_total_amount',
        'Flipkart Order ID': 'marketplace_order_ref',
        'Flipkart Order Date': 'marketplace_order_date',
        'Flipkart Payment Status': 'marketplace_payment_status',
        'Flipkart Order Status': 'marketplace_order_status',
        'Flipkart Total Amount': 'marketplace_total_amount',
        'Amazon Order ID': 'marketplace_order_ref',
        'Amazon Order Date': 'marketplace_order_date',
        'Amazon Payment Status': 'marketplace_payment_status',
        'Amazon Order Status': 'marketplace_order_status',
        'Amazon Total Amount': 'marketplace_total_amount',
        'Blinkit Order ID': 'marketplace_order_ref',
        'Blinkit Order Date': 'marketplace_order_date',
        'Blinkit Delivery Slot': 'marketplace_delivery_slot',
        'Blinkit Payment Status': 'marketplace_payment_status',
        'Blinkit Order Status': 'marketplace_order_status',
        'Blinkit Total Amount': 'marketplace_total_amount',
        'Shopify Order ID': 'marketplace_order_ref',
        'Shopify Order Date': 'marketplace_order_date',
        'Shopify Payment Status': 'marketplace_payment_status',
        'Shopify Order Status': 'marketplace_order_status',
        'Shopify Total Amount': 'marketplace_total_amount',
        'Invoice Number': 'marketplace_invoice_number',
        'Invoice Type': 'marketplace_invoice_type',
        'State of Sale': 'marketplace_sale_state',
        'Payment Terms': 'payment_term_name',
    }

    @api.onchange('xlsx_file')
    def _onchange_xlsx_file(self):
        if not self.xlsx_file:
            self.report_file = False
            self.report_filename = False

    def _validate_row(self, row, index):
        errors = []
        try:
            # Validate marketplace configuration presence for this specific row
            m_type = str(row.get('marketplace_type_excel') or '').strip()
            if not m_type:
                errors.append('Missing Marketplace Type')
            else:
                m_rec = self.env['marketplace.master'].search([
                    '|', ('name', '=ilike', m_type), ('code', '=ilike', m_type)
                ], limit=1)
                if not m_rec:
                    errors.append(f"Marketplace '{m_type}' not found in system configurations")

            errors += self._validate_basic_fields(row)
            errors += self._validate_customer_fields(row)
            errors += self._validate_address_fields(row)
            errors += self._validate_product_fields(row)
            errors += self._validate_pricing_fields(row)
            errors += self._validate_misc_fields(row)
            errors += self._validate_csv_injection(row)
        except Exception as e:
            errors.append(f"Unexpected error in row {index}: {str(e)}")
        return '; '.join(filter(None, errors))

    def _validate_basic_fields(self, row):
        errors = []
        if not row.get('marketplace_order_id'):
            errors.append('Missing marketplace_order_id')

        order_date = row.get('order_date')
        if not order_date:
            errors.append('Missing order_date')
        elif not isinstance(order_date, (datetime, str)):
            errors.append('order_date must be a valid date or string format')
        elif isinstance(order_date, str):
            try:
                datetime.fromisoformat(order_date.replace('Z', ''))
            except ValueError:
                try:
                    datetime.strptime(order_date, '%Y-%m-%d')
                except ValueError:
                    errors.append('order_date must be in YYYY-MM-DD format')
        return errors

    def _validate_customer_fields(self, row):
        errors = []
        if not row.get('customer_name'):
            errors.append('Missing customer_name')
        customer_email = row.get('customer_email')
        if not customer_email:
            errors.append('Missing customer_email')
        elif isinstance(customer_email, str) and '@' not in customer_email:
            errors.append('Invalid email format')
        return errors

    def _validate_address_fields(self, row):
        errors = []
        for field in ['customer_street', 'customer_city', 'customer_state']:
            if not row.get(field):
                errors.append(f"Missing {field}")
        customer_zip = row.get('customer_zip')
        if not customer_zip:
            errors.append('Missing customer_zip')
        elif not str(customer_zip).strip().isdigit():
            errors.append('customer_zip must be numeric')
        return errors

    def _validate_product_fields(self, row):
        errors = []
        product_sku = str(row.get('product_sku') or '').strip()
        if not product_sku:
            errors.append('Missing product_sku')
        else:
            product = self.env['product.product'].search([
                '|', ('default_code', '=', product_sku), ('barcode', '=', product_sku)
            ], limit=1)
            if not product:
                errors.append(f"Product with SKU/Barcode '{product_sku}' not found")

        quantity = row.get('quantity')
        if quantity in (None, ''):
            errors.append('Missing quantity')
        else:
            try:
                if float(quantity) <= 0:
                    errors.append('quantity must be greater than 0')
            except ValueError:
                errors.append('quantity must be a number')
        return errors

    def _validate_pricing_fields(self, row):
        errors = []
        for field in ['unit_price', 'tax_percent', 'shipping_amount']:
            val = row.get(field)
            if val not in (None, ''):
                try:
                    if float(val) < 0:
                        errors.append(f"{field} cannot be negative")
                except ValueError:
                    errors.append(f"{field} must be a number")
        return errors

    def _validate_misc_fields(self, row):
        errors = []
        if not row.get('currency'):
            errors.append('Missing currency')
        return errors

    def _validate_csv_injection(self, row):
        errors = []
        dangerous_prefixes = ('=', '+', '-', '@')
        for key, value in row.items():
            if isinstance(value, str) and value.startswith(dangerous_prefixes):
                errors.append(f"Incorrect injection in field '{key}'")
        return errors

    def action_import_orders(self):
        self.ensure_one()
        try:
            headers, data_rows = self._read_xlsx()
        except Exception as e:
            raise UserError(f"Failed to read XLSX file: {str(e)}")

        missing = self._validate_headers(headers, self.DEFAULT_EXPECTED_HEADERS)
        if missing:
            raise UserError(f"The following required columns are missing or incorrect: {', '.join(missing)}")

        if self.log_ids:
            self.log_ids.unlink()

        headers_list = list(headers)
        if 'Status' not in headers_list:
            headers_list.append('Status')
        if 'Error' not in headers_list:
            headers_list.append('Error')

        parsed_rows = []
        has_error = False

        for idx, raw in enumerate(data_rows, start=1):
            row = {k: (v if v is not None else '') for k, v in raw.items()}
            row_error = self._validate_row(row, idx)
            row['Error'] = row_error

            if row_error:
                row['Status'] = 'Failed'
                has_error = True
                self._log_failure(idx, row_error)
            else:
                row['Status'] = 'Success'
            parsed_rows.append(row)

        if has_error:
            self._generate_report_file(headers_list, parsed_rows)
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/?model=marketplace.order.import.wizard&field=report_file&id={self.id}&filename={self.report_filename}&download=true',
                'target': 'new',
            }

        grouped = self._group_rows_by_order(parsed_rows)
        created_orders = []
        failed_rows = []

        # Process each group using its dedicated row-level marketplace record context
        for key_tuple, group in grouped.items():
            marketplace_order_id, marketplace_record = key_tuple
            try:
                with self.env.cr.savepoint():
                    order_result = self._process_order_group(marketplace_order_id, marketplace_record, group)
                    if order_result.get('success'):
                        created_orders.append(order_result.get('order'))
                    else:
                        failed_rows.extend(order_result.get('failed_rows', []))
            except Exception as e:
                err_msg = f"Unexpected processing error: {str(e)}"
                for idx, _ in group:
                    self._log_failure(idx, err_msg)
                    failed_rows.append((idx, err_msg))

        final_rows = self._attach_log_messages_to_rows(parsed_rows)
        self._generate_report_file(headers_list, final_rows, created_orders=created_orders)

        if failed_rows:
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/?model=marketplace.order.import.wizard&field=report_file&id={self.id}&filename={self.report_filename}&download=true',
                'target': 'new',
            }

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
            'params': {
                'message': f"Successfully imported mixed marketplace entries! Created {len(created_orders)} orders."}
        }

    def _read_xlsx(self):
        if not self.xlsx_file:
            raise UserError("Please upload an XLSX file.")
        if not (self.xlsx_filename or "").lower().endswith('.xlsx'):
            raise UserError("Only XLSX files are allowed.")

        file_content = base64.b64decode(self.xlsx_file)
        workbook = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise UserError("XLSX file is empty or invalid.")

        raw_headers = [str(h).strip() for h in rows[0] if h is not None]
        headers = [self.HEADER_MAP.get(h, h.strip()) for h in raw_headers]

        data_rows = []
        for r in rows[1:]:
            if not any(cell is not None for cell in r):
                continue
            row_dict = dict(zip(headers, r))
            row_dict = {k: (v.strip() if isinstance(v, str) else v) for k, v in row_dict.items() if k}
            data_rows.append(row_dict)

        return headers, data_rows

    def _validate_headers(self, headers, expected):
        lower_headers = [str(h).lower() for h in headers]
        return [h for h in expected if h.lower() not in lower_headers]

    # --- CHANGED: Group by BOTH Order ID and Row Marketplace Record ---
    def _group_rows_by_order(self, parsed_rows):
        groups = {}
        for idx, row in enumerate(parsed_rows, start=1):
            mid = str(row.get('marketplace_order_id') or '').strip()
            m_type = str(row.get('marketplace_type_excel') or '').strip()

            # Dynamically look up the unique marketplace record for this row
            m_record = self.env['marketplace.master'].search([
                '|', ('name', '=ilike', m_type), ('code', '=ilike', m_type)
            ], limit=1)

            # Use tuple key: (Order_ID, Marketplace_Record)
            key = (mid, m_record)
            if key not in groups:
                groups[key] = []
            groups[key].append((idx, row))
        return groups

    def _find_existing_marketplace_order(self, marketplace_order_id, marketplace_record):
        if not marketplace_order_id or not marketplace_record:
            return self.env['sale.order']
        origin = f"marketplace:{marketplace_record.name}|{marketplace_order_id}"
        return self.env['sale.order'].search([
            '|', ('origin', '=', origin), ('marketplace_order_ref', '=', marketplace_order_id)
        ], limit=1)

    # --- CHANGED: Explicitly receives individual row-level marketplace context record ---
    def _process_order_group(self, marketplace_order_id, marketplace_record, rows):
        failed_rows = []
        first_row = rows[0][1]

        customer = self._resolve_customer(first_row)
        if not customer:
            msg = 'Customer not found and automation creation is disabled.'
            for idx, _ in rows:
                self._log_failure(idx, msg)
            return {'success': False, 'failed_rows': [(idx, msg) for idx, _ in rows]}

        marketplace_type = marketplace_record.code or 'unknown'
        existing_order = self._find_existing_marketplace_order(marketplace_order_id, marketplace_record)
        if existing_order:
            msg = f"Sale order {existing_order.name} already exists for {marketplace_record.name}."
            for idx, _ in rows:
                self._log_failure(idx, msg)
            return {'success': False, 'failed_rows': [(idx, msg) for idx, _ in rows]}

        origin = f"marketplace:{marketplace_record.name}|{marketplace_order_id}"
        warehouse = getattr(marketplace_record, 'warehouse_map', self.env.ref('stock.warehouse0'))

        order_vals = {
            'partner_id': customer.id,
            'date_order': first_row.get('order_date') or fields.Datetime.now(),
            'origin': origin,
            'warehouse_id': warehouse.id if warehouse else self.env.ref('stock.warehouse0').id,
            'team_id': self.env.ref('sales_team.salesteam_website_sales').id,
            'marketplace_type': marketplace_type,
            'marketplace_order_ref': marketplace_order_id,
        }

        payment_term_name = first_row.get('payment_term_name')
        if payment_term_name:
            term = self.env['account.payment.term'].search([('name', '=ilike', payment_term_name.strip())], limit=1)
            if term:
                order_vals['payment_term_id'] = term.id

        order_tag_name = str(first_row.get('order_tag') or '').strip() or marketplace_record.name
        if order_tag_name:
            tag = self.env['crm.tag'].search([('name', '=', order_tag_name)], limit=1)
            if not tag:
                tag = self.env['crm.tag'].create({'name': order_tag_name})
            order_vals['tag_ids'] = [(4, tag.id)]

        marketplace_fields = [
            'marketplace_order_date', 'marketplace_payment_status', 'marketplace_order_status',
            'marketplace_delivery_slot', 'marketplace_total_amount', 'marketplace_invoice_number',
            'marketplace_invoice_type', 'marketplace_sale_state', 'marketplace_hsn_code', 'marketplace_item_id'
        ]
        for fld in marketplace_fields:
            if first_row.get(fld):
                order_vals[fld] = first_row.get(fld)

        sale_order = self.env['sale.order'].create(order_vals)

        for idx, row in rows:
            sku = str(row.get('product_sku') or '').strip()
            product = self.env['product.product'].search([
                '|', ('default_code', '=', sku), ('barcode', '=', sku)
            ], limit=1)

            if not product:
                msg = f"Product SKU {sku} extraction failed."
                self._log_failure(idx, msg)
                failed_rows.append((idx, msg))
                continue

            row_hsn = str(row.get('marketplace_hsn_code') or '').strip()
            if row_hsn:
                hsn_field = 'l10n_in_hsn_code'
                if hasattr(product, hsn_field) and not product[hsn_field]:
                    product.product_tmpl_id.write({hsn_field: row_hsn})

            qty = float(row.get('quantity') or 1.0)
            unit_price = float(row.get('unit_price')) if row.get('unit_price') not in (None, '') else product.lst_price

            line_vals = {
                'order_id': sale_order.id,
                'product_id': product.id,
                'product_uom_qty': qty,
                'price_unit': unit_price,
                'name': row.get('order_line_note') or product.name,
            }

            tax_percent = row.get('tax_percent')
            if tax_percent not in (None, ''):
                try:
                    t_pct = float(tax_percent)
                    tax = self.env['account.tax'].search([('amount', '=', t_pct), ('type_tax_use', '=', 'sale')],
                                                         limit=1)
                    if not tax:
                        tax = self.env['account.tax'].create({
                            'name': f"{int(t_pct)}% VAT",
                            'amount': t_pct,
                            'type_tax_use': 'sale',
                            'amount_type': 'percent',
                        })
                    line_vals['tax_id'] = [(6, 0, [tax.id])]
                except Exception:
                    pass

            self.env['sale.order.line'].create(line_vals)

        if self.confirm_orders and not failed_rows:
            try:
                sale_order.action_confirm()
            except Exception as e:
                return {'success': False, 'order': sale_order, 'failed_rows': [(idx, str(e)) for idx, _ in rows]}

        return {'success': True, 'order': sale_order, 'failed_rows': failed_rows}

    def _resolve_country(self, row):
        country_raw = row.get('customer_country') or 'India'
        country_str = str(country_raw).strip()
        return self.env['res.country'].search(['|', ('code', '=ilike', country_str), ('name', '=ilike', country_str)],
                                              limit=1)

    def _resolve_state(self, row, country_id):
        state_raw = row.get('customer_state') or row.get('supply_state')
        if not state_raw or not country_id:
            return self.env['res.country.state']
        state_str = str(state_raw).strip()
        return self.env['res.country.state'].search(
            [('country_id', '=', country_id.id), '|', ('code', '=ilike', state_str), ('name', '=ilike', state_str)],
            limit=1)

    def _resolve_customer(self, first_row):
        email = str(first_row.get('customer_email') or '').strip()
        name = str(first_row.get('customer_name') or '').strip()

        customer = None
        if email:
            customer = self.env['res.partner'].search([('email', '=', email)], limit=1)
        if not customer and name:
            customer = self.env['res.partner'].search([('name', '=', name)], limit=1)

        if not customer and self.create_customer_if_missing:
            country_record = self._resolve_country(first_row)
            state_record = self._resolve_state(first_row, country_record)

            partner_vals = {
                'name': name or email or 'Unknown Customer',
                'email': email or False,
                'street': first_row.get('customer_street') or False,
                'city': first_row.get('customer_city') or False,
                'zip': first_row.get('customer_zip') or False,
                'country_id': country_record.id if country_record else False,
                'state_id': state_record.id if state_record else False,
                'vat': str(first_row.get('vat')).strip() if first_row.get('vat') else False,
            }
            customer = self.env['res.partner'].create(partner_vals)
        return customer

    def _log_failure(self, row_index, message):
        self.env['marketplace.import.log'].create({
            'wizard_id': self.id,
            'row_index': row_index,
            'status': 'failed',
            'message': str(message),
        })

    def _generate_report_file(self, headers, rows, created_orders=None):
        created_orders = created_orders or []
        output = io.BytesIO()
        workbook_out = xlsxwriter.Workbook(output, {'in_memory': True})

        header_format = workbook_out.add_format(
            {'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1, 'align': 'center'})
        success_format = workbook_out.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'border': 1})
        failed_format = workbook_out.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'border': 1})
        default_format = workbook_out.add_format({'border': 1})
        date_format = workbook_out.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})

        worksheet = workbook_out.add_worksheet('Import Report')
        worksheet.freeze_panes(1, 0)

        headers_sorted = [h for h in headers if str(h).lower() not in ('status', 'error')] + ['Status', 'Error']
        col_widths = [max(len(str(h)), 15) for h in headers_sorted]

        for col_num, header in enumerate(headers_sorted):
            display_name = str(header).replace('_', ' ').title()
            worksheet.write(0, col_num, display_name, header_format)

        for row_num, row_data in enumerate(rows, start=1):
            row_dict = dict(row_data)
            status_val = str(row_dict.get('Status') or row_dict.get('status') or 'Success')
            error_val = str(row_dict.get('Error') or row_dict.get('error') or '')

            for col_num, column in enumerate(headers_sorted):
                value = row_dict.get(column, '')
                cell_format = default_format

                if str(column).lower() == 'status':
                    value = status_val
                    cell_format = success_format if status_val.lower() == 'success' else failed_format
                elif str(column).lower() == 'error':
                    value = error_val
                    cell_format = failed_format if error_val else success_format
                elif str(column).lower() == 'order_date' and value:
                    if isinstance(value, (datetime, fields.Datetime)):
                        cell_format = date_format
                    else:
                        try:
                            value = datetime.fromisoformat(str(value).replace('Z', ''))
                            cell_format = date_format
                        except ValueError:
                            pass

                if cell_format == date_format and isinstance(value, datetime):
                    worksheet.write_datetime(row_num, col_num, value, cell_format)
                else:
                    worksheet.write(row_num, col_num, str(value) if value is not None else '', cell_format)

                col_widths[col_num] = max(col_widths[col_num], len(str(value or '')) + 2)

        for i, width in enumerate(col_widths):
            worksheet.set_column(i, i, min(width, 50))

        ws_sum = workbook_out.add_worksheet('Summary')
        summary_header = workbook_out.add_format(
            {'bold': True, 'bg_color': '#4BACC6', 'font_color': 'white', 'border': 1})
        summary_value = workbook_out.add_format({'border': 1})

        ws_sum.write(0, 0, 'Total Rows', summary_header)
        ws_sum.write(0, 1, len(rows), summary_value)
        ws_sum.write(1, 0, 'Created Orders', summary_header)
        ws_sum.write(1, 1, len(created_orders), summary_value)

        failed_count = self.env['marketplace.import.log'].search_count(
            [('wizard_id', '=', self.id), ('status', '=', 'failed')])
        ws_sum.write(2, 0, 'Failed Rows', summary_header)
        ws_sum.write(2, 1, failed_count, summary_value)

        workbook_out.close()
        output.seek(0)
        self.report_file = base64.b64encode(output.read())

        current_user_dt = fields.Datetime.context_timestamp(self, fields.Datetime.now())
        self.report_filename = f"Import_Report_{current_user_dt.strftime('%d-%m-%Y_%I.%M_%p')}.xlsx"

    def _attach_log_messages_to_rows(self, parsed_rows):
        logs = self.env['marketplace.import.log'].search([('wizard_id', '=', self.id)])
        log_map = {}
        for l in logs:
            log_map.setdefault(int(l.row_index), []).append(l.message)

        final_rows = []
        for idx, row in enumerate(parsed_rows, start=1):
            row_copy = dict(row)
            messages = log_map.get(idx, [])

            if messages:
                row_copy['Error'] = '; '.join(messages)
                row_copy['Status'] = 'Failed'
            else:
                row_copy['Error'] = ''
                row_copy['Status'] = 'Success'
            final_rows.append(row_copy)
        return final_rows