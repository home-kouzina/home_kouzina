# -*- coding: utf-8 -*-
import base64
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError


class InventoryVariation(models.Model):
    """
    Extends inventory.variation with:
      - Download Sample Template  (header button)
      - Upload File field
      - Process Inventory Upload  (header button)
      - Download Error Report     (header button)
      - Upload Logs tab
    """
    _inherit = 'inventory.variation'

    # ------------------------------------------------------------------
    # Upload fields added to the existing model
    # ------------------------------------------------------------------

    upload_file = fields.Binary(string='Excel File', attachment=True)
    upload_filename = fields.Char(string='File Name')
    upload_date = fields.Datetime(string='Upload Date', readonly=True)
    uploaded_by = fields.Many2one('res.users', string='Uploaded By', readonly=True)
    total_rows = fields.Integer(string='Total Rows', readonly=True)
    success_rows = fields.Integer(string='Success Rows', readonly=True)
    failed_rows = fields.Integer(string='Failed Rows', readonly=True)
    message_summary = fields.Text(string='Upload Summary', readonly=True)
    error_file = fields.Binary(string='Error Report', attachment=True, readonly=True)
    error_filename = fields.Char(string='Error File Name', readonly=True)
    upload_line_ids = fields.One2many(
        'inventory.variation.upload.line', 'variation_id',
        string='Upload Logs', readonly=True,
    )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _normalize(self, value):
        return str(value).strip() if value is not None else ''

    def _get_binary_download_url(self, field_name, filename_field):
        self.ensure_one()
        return (
            '/web/content?model=%s&id=%s&field=%s&filename_field=%s&download=true'
            % (self._name, self.id, field_name, filename_field)
        )

    # ------------------------------------------------------------------
    # Template download
    # ------------------------------------------------------------------

    def action_download_sample_template(self):
        """
        Generate and download an Excel template pre-filled from existing
        variation lines (after Load Products) or from stock quants.

        Columns the user sees and fills:
            Product | Product Type | Sale Price | Location | UoM | Physical Qty

        System Qty and Variation are intentionally excluded — they are
        always fetched/computed server-side after upload.
        """
        self.ensure_one()

        if self.line_ids:
            rows = self._template_rows_from_variation_lines()
        else:
            rows = self._template_rows_from_stock()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Variation Upload'

        header_font   = Font(bold=True, color='FFFFFF', size=11)
        header_fill   = PatternFill('solid', fgColor='1F4E78')
        header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin          = Side(style='thin')
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)
        locked_fill   = PatternFill('solid', fgColor='D9E1F2')  # pale blue  – info only
        editable_fill = PatternFill('solid', fgColor='E2EFDA')  # pale green – user fills

        # Visible columns + 2 hidden key columns at the end (product_id, location_id)
        headers = ['Product', 'Product Type', 'Sale Price', 'Location', 'UoM', 'Physical Qty',
                   'product_id', 'location_id']
        HIDDEN_FROM = 7  # columns 7 & 8 are hidden key columns (1-indexed)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if col_idx < HIDDEN_FROM:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = border

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx < HIDDEN_FROM:
                    cell.border    = border
                    cell.alignment = Alignment(vertical='center')
                    cell.fill = editable_fill if col_idx == 6 else locked_fill

        # Hide the key columns from the user
        ws.column_dimensions[get_column_letter(HIDDEN_FROM)].hidden = True
        ws.column_dimensions[get_column_letter(HIDDEN_FROM + 1)].hidden = True

        widths = [45, 18, 14, 40, 12, 15]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        file_content = base64.b64encode(output.getvalue())

        filename = 'inventory_variation_template_%s.xlsx' % self.name
        self.write({'error_file': file_content, 'error_filename': filename})

        return {
            'type': 'ir.actions.act_url',
            'url': self._get_binary_download_url('error_file', 'error_filename'),
            'target': 'self',
        }

    def _product_display_name(self, product):
        """
        Return the full display name of a product.product record.
        Uses display_name which includes variant attribute values
        (e.g. 'Coriander seeds (green extra)') so the template
        exactly matches what Odoo shows on screen.
        """
        return product.display_name or product.name

    def _template_rows_from_variation_lines(self):
        """Build template rows from existing variation lines on this record."""
        rows = []
        for line in self.line_ids:
            product  = line.product_id
            location = line.location_id
            rows.append([
                self._product_display_name(product),                                    # A – display
                line.product_type or ('Finished' if product.is_finished_good else 'Raw Material'),
                line.sale_price or product.lst_price,
                location.complete_name if location else '',
                line.uom_id.name if line.uom_id else (product.uom_id.name if product.uom_id else ''),
                0.0,          # F – Physical Qty – filled by user
                product.id,   # G – hidden: product.product id for exact matching
                location.id if location else '',  # H – hidden: location id
            ])
        return rows

    def _template_rows_from_stock(self):
        """Fallback: build rows from stock quants for the selected warehouses."""
        if self.warehouse_ids:
            root_locations = self.warehouse_ids.mapped('lot_stock_id')
            domain = [('id', 'child_of', root_locations.ids), ('usage', '=', 'internal')]
            locations = self.env['stock.location'].search(domain)
        else:
            locations = self.env['stock.location'].search([('usage', '=', 'internal')])

        quants = self.env['stock.quant'].search([
            ('location_id', 'in', locations.ids),
            ('product_id.active', '=', True),
        ])
        rows = []
        for q in quants:
            product  = q.product_id
            location = q.location_id
            rows.append([
                self._product_display_name(product),
                'Finished' if product.is_finished_good else 'Raw Material',
                product.lst_price,
                location.complete_name,
                product.uom_id.name,
                0.0,
                product.id,   # hidden key
                location.id,  # hidden key
            ])
        return rows

    # ------------------------------------------------------------------
    # Upload processing
    # ------------------------------------------------------------------

    def action_process_inventory_upload(self):
        """
        Parse the uploaded Excel file and set Physical Qty on variation lines.
        System Qty is fetched from stock.quant; Variation is recomputed.
        """
        self.ensure_one()
        if self.state != 'draft':
            raise UserError(_('Upload is only allowed on Draft variation sessions.'))
        if not self.upload_file:
            raise UserError(_('Please attach an Excel file before processing.'))

        # Clear previous upload logs
        self.upload_line_ids.unlink()
        self.write({
            'error_file': False,
            'error_filename': False,
            'message_summary': False,
            'total_rows': 0,
            'success_rows': 0,
            'failed_rows': 0,
        })

        # Parse workbook
        try:
            wb = load_workbook(filename=io.BytesIO(base64.b64decode(self.upload_file)), data_only=True)
        except Exception as e:
            raise UserError(_('Could not read the Excel file: %s') % str(e))

        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('The uploaded file is empty.'))

        headers = [self._normalize(h) for h in rows[0]]
        expected = ['Product', 'Product Type', 'Sale Price', 'Location', 'UoM', 'Physical Qty']
        if headers[:6] != expected:
            raise UserError(_(
                'Invalid template. Expected columns:\n%s\n\nFound:\n%s'
            ) % (', '.join(expected), ', '.join(headers[:6])))

        error_rows   = []
        success_rows = 0
        failed_rows  = 0
        total_rows   = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(c in (None, '') for c in row[:6]):
                continue

            total_rows += 1
            product_raw      = self._normalize(row[0] if len(row) > 0 else '')
            location_raw     = self._normalize(row[3] if len(row) > 3 else '')
            physical_qty_raw = row[5] if len(row) > 5 else None

            # Hidden key columns written by the template generator (col G=index6, col H=index7)
            hidden_product_id  = row[6] if len(row) > 6 else None
            hidden_location_id = row[7] if len(row) > 7 else None

            # Validate qty
            if physical_qty_raw in (None, ''):
                physical_qty_raw = 0.0
            try:
                physical_qty = float(physical_qty_raw)
            except (TypeError, ValueError):
                failed_rows += 1
                error_rows.append([row_idx, product_raw, location_raw, physical_qty_raw,
                                    _('Physical Qty must be numeric.')])
                self._create_upload_log(row_idx, product_raw, location_raw,
                                        physical_qty_raw, 'failed', _('Physical Qty must be numeric.'))
                continue

            if physical_qty < 0:
                failed_rows += 1
                error_rows.append([row_idx, product_raw, location_raw, physical_qty,
                                    _('Physical Qty cannot be negative.')])
                self._create_upload_log(row_idx, product_raw, location_raw,
                                        physical_qty, 'failed', _('Physical Qty cannot be negative.'))
                continue

            # ----------------------------------------------------------
            # Resolve product — prefer hidden ID (exact, variant-safe),
            # fall back to name matching for manually-typed templates
            # ----------------------------------------------------------
            product = None
            if hidden_product_id:
                try:
                    pid = int(hidden_product_id)
                    product = self.env['product.product'].browse(pid)
                    if not product.exists():
                        product = None
                except (TypeError, ValueError):
                    product = None

            if not product:
                try:
                    product = self._resolve_upload_product(product_raw)
                except ValidationError as e:
                    msg = str(e.args[0]) if e.args else str(e)
                    failed_rows += 1
                    error_rows.append([row_idx, product_raw, location_raw, physical_qty, msg])
                    self._create_upload_log(row_idx, product_raw, location_raw,
                                            physical_qty, 'failed', msg)
                    continue

            # ----------------------------------------------------------
            # Resolve location — prefer hidden ID, fall back to name
            # ----------------------------------------------------------
            location = None
            if hidden_location_id:
                try:
                    lid = int(hidden_location_id)
                    location = self.env['stock.location'].browse(lid)
                    if not location.exists():
                        location = None
                except (TypeError, ValueError):
                    location = None

            if not location:
                try:
                    location = self._resolve_upload_location(location_raw)
                except ValidationError as e:
                    msg = str(e.args[0]) if e.args else str(e)
                    failed_rows += 1
                    error_rows.append([row_idx, product_raw, location_raw, physical_qty, msg])
                    self._create_upload_log(row_idx, product_raw, location_raw,
                                            physical_qty, 'failed', msg)
                    continue

            # Update / create variation line
            try:
                line     = self._get_or_create_variation_line(product, location)
                old_phys = line.physical_qty
                line.write({'physical_qty': physical_qty})

                success_rows += 1
                self._create_upload_log(
                    row_idx,
                    self._product_display_name(product),
                    location.complete_name,
                    physical_qty,
                    'success',
                    _('Physical Qty updated from %.3f to %.3f.') % (old_phys, physical_qty),
                    product_id=product.id, location_id=location.id,
                    old_qty=old_phys, new_qty=physical_qty,
                )
            except Exception as e:
                failed_rows += 1
                msg = str(e)
                error_rows.append([row_idx, product_raw, location_raw, physical_qty, msg])
                self._create_upload_log(row_idx, product_raw, location_raw,
                                        physical_qty, 'failed', msg)

        if error_rows:
            self._generate_upload_error_report(error_rows)

        summary = _('%d rows processed: %d succeeded, %d failed.') % (
            total_rows, success_rows, failed_rows)

        self.write({
            'upload_date':      fields.Datetime.now(),
            'uploaded_by':      self.env.user.id,
            'total_rows':       total_rows,
            'success_rows':     success_rows,
            'failed_rows':      failed_rows,
            'message_summary':  summary,
        })

        return {
            'type': 'ir.actions.client',
            'tag':  'display_notification',
            'params': {
                'title':   _('Upload Complete'),
                'message': summary,
                'sticky':  False,
                'type':    'success' if not failed_rows else 'warning',
                'next':    {'type': 'ir.actions.client', 'tag': 'reload'},
            },
        }

    # ------------------------------------------------------------------
    # Resolution helpers
    # ------------------------------------------------------------------

    def _resolve_upload_product(self, product_raw):
        if not product_raw:
            raise ValidationError(_('Product cell is empty.'))
        Product = self.env['product.product']

        # [CODE] Name
        if product_raw.startswith('[') and '] ' in product_raw:
            code = product_raw[1:].split('] ', 1)[0].strip()
            p = Product.search([('default_code', '=', code), ('active', '=', True)], limit=2)
            if len(p) > 1:
                raise ValidationError(_('Multiple products for reference "%s".') % code)
            if p:
                return p

        # Exact name
        p = Product.search([('name', '=', product_raw), ('active', '=', True)], limit=2)
        if len(p) > 1:
            raise ValidationError(_('Multiple products for name "%s". Use [CODE] Name format.') % product_raw)
        if p:
            return p

        # Bare code
        p = Product.search([('default_code', '=', product_raw), ('active', '=', True)], limit=2)
        if len(p) > 1:
            raise ValidationError(_('Multiple products for reference "%s".') % product_raw)
        if p:
            return p

        raise ValidationError(_('Product not found: %s') % product_raw)

    def _resolve_upload_location(self, location_raw):
        if not location_raw:
            raise ValidationError(_('Location cell is empty.'))
        Location = self.env['stock.location']
        base = [('usage', 'in', ('internal', 'transit'))]

        for domain_field in ['complete_name', 'barcode', 'name']:
            loc = Location.search(base + [(domain_field, '=', location_raw)], limit=2)
            if len(loc) > 1:
                raise ValidationError(
                    _('Multiple locations for "%s". Use the full location path.') % location_raw)
            if loc:
                return loc

        raise ValidationError(_('Location not found: %s') % location_raw)

    def _get_or_create_variation_line(self, product, location):
        """Return existing line or create a new one with system qty from quants."""
        line = self.line_ids.filtered(
            lambda l: l.product_id.id == product.id and l.location_id.id == location.id
        )
        if line:
            return line[0]

        quant = self.env['stock.quant'].search([
            ('product_id', '=', product.id),
            ('location_id', '=', location.id),
        ], limit=1)
        theoretical_qty = quant.quantity if quant else 0.0
        product_type    = 'Finished' if product.is_finished_good else 'Raw Material'

        return self.env['inventory.variation.line'].create({
            'variation_id':   self.id,
            'product_id':     product.id,
            'location_id':    location.id,
            'uom_id':         product.uom_id.id,
            'theoretical_qty': theoretical_qty,
            'physical_qty':   0.0,
            'product_type':   product_type,
            'sale_price':     product.lst_price,
        })

    # ------------------------------------------------------------------
    # Upload log helper
    # ------------------------------------------------------------------

    def _create_upload_log(self, row_number, product_name, location_name,
                           new_qty, status, message,
                           product_id=False, location_id=False,
                           old_qty=0.0):
        self.env['inventory.variation.upload.line'].create({
            'variation_id':  self.id,
            'row_number':    row_number,
            'product_name':  product_name,
            'location_name': location_name,
            'product_id':    product_id,
            'location_id':   location_id,
            'old_qty':       old_qty,
            'new_qty':       float(new_qty) if isinstance(new_qty, (int, float)) else 0.0,
            'status':        status,
            'message':       message,
        })

    # ------------------------------------------------------------------
    # Error report
    # ------------------------------------------------------------------

    def _generate_upload_error_report(self, error_rows):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Error Report'
        ws.append(['Excel Row', 'Product', 'Location', 'Physical Qty', 'Error Message'])
        for row in error_rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        self.write({
            'error_file':    base64.b64encode(output.getvalue()),
            'error_filename': '%s_error_report.xlsx' % self.name,
        })

    def action_download_error_report(self):
        self.ensure_one()
        if not self.error_file:
            raise UserError(_('No error report available.'))
        return {
            'type': 'ir.actions.act_url',
            'url':  self._get_binary_download_url('error_file', 'error_filename'),
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Override Load Products to pre-fill physical_qty from last IVR
    # ------------------------------------------------------------------

    def action_load_products(self):
        """
        Calls the original Load Products, then back-fills physical_qty on
        every line from the most recently confirmed IVR session that has a
        line for the same (product, location).

        This means when a user opens a new IVR and clicks Load Products,
        the Physical Qty column already shows the last recorded physical
        count instead of 0.
        """
        # Run the original logic first (builds all lines with physical_qty=0)
        result = super().action_load_products()

        # Build a lookup: (product_id, location_id) -> physical_qty
        # from the most recently confirmed IVR (excluding the current one)
        last_physical = self._get_last_confirmed_physical_qtys()

        if last_physical:
            for line in self.line_ids:
                key = (line.product_id.id, line.location_id.id)
                if key in last_physical:
                    line.physical_qty = last_physical[key]

            # Recompute variation after updating physical qtys
            self.line_ids._compute_variation_qty()

        return result

    def _get_last_confirmed_physical_qtys(self):
        """
        Returns a dict of {(product_id, location_id): physical_qty} from
        the most recent confirmed/reported IVR session (other than self)
        that shares at least one of the same warehouses.
        """
        # Find the latest confirmed IVR for the same warehouses
        domain = [
            ('id', '!=', self.id),
            ('state', 'in', ('confirmed', 'reported')),
        ]
        if self.warehouse_ids:
            domain += [('warehouse_ids', 'in', self.warehouse_ids.ids)]

        last_ivr = self.env['inventory.variation'].search(
            domain,
            order='id desc',
            limit=1,
        )

        if not last_ivr:
            return {}

        return {
            (line.product_id.id, line.location_id.id): line.physical_qty
            for line in last_ivr.line_ids
            if line.physical_qty != 0.0
        }


# ---------------------------------------------------------------------------
# Upload log line  (linked directly to inventory.variation)
# ---------------------------------------------------------------------------

class InventoryVariationUploadLine(models.Model):
    _name        = 'inventory.variation.upload.line'
    _description = 'Inventory Variation Upload Log Line'
    _order       = 'id asc'

    variation_id  = fields.Many2one('inventory.variation', string='Variation',
                                    required=True, ondelete='cascade')
    row_number    = fields.Integer(string='Excel Row')
    product_name  = fields.Char(string='Product')
    location_name = fields.Char(string='Location')
    product_id    = fields.Many2one('product.product', string='Matched Product')
    location_id   = fields.Many2one('stock.location',  string='Matched Location')
    old_qty       = fields.Float(string='Old Physical Qty', digits='Product Unit of Measure')
    new_qty       = fields.Float(string='New Physical Qty',  digits='Product Unit of Measure')
    status        = fields.Selection([
        ('success', 'Success'),
        ('failed',  'Failed'),
    ], string='Status', default='success')
    message       = fields.Text(string='Message')
